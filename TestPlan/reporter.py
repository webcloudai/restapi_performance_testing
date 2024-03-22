'''
© 2024 Daniil Sokolov <daniil.sokolov@webcloudai.com>
MIT License
'''
from abc import ABC, abstractmethod
import dataclasses
from enum import Enum
from typing import List, Union, Tuple, Sequence
from pathlib import Path
import os
from uuid import uuid4
import json
import datetime
import openpyxl
import logging
_top_logger = logging.getLogger(__name__)

class LogRecordType(Enum):
    ERROR = "error"
    LATENCY = "latency"
    OTHER = "other"

@dataclasses.dataclass
class LogRecord:
    stage:str
    job:str
    task:str
    logType:LogRecordType
    data:Union[dict, list, str]

    def as_dict(self)->dict:
        return {
            "stage": self.stage,
            "job": self.job,
            "task": self.task,
            "logType": self.logType.value if isinstance(self.logType, LogRecordType) else self.logType,
            "data": self.data
        }

class Reporter(ABC):
    ''' '''
    def __init__(self, reporter_options:dict):
        ''' '''
    
    @abstractmethod
    def add(self, record:LogRecord):
        ''' '''
    @abstractmethod
    def add_bunch(self, records:List[LogRecord]):
        ''' '''
    @abstractmethod
    def list_all(self, record_type:LogRecordType)->List[str]:
        ''' '''
    @abstractmethod
    def get_all(self, record_type:LogRecordType)->List[LogRecord]:
        ''' '''
    @abstractmethod
    def get_one(self, record_id)->LogRecord:
        ''' '''


class ReporterJsonRecords(Reporter):
    ''' '''
    def __init__(self, reporter_options:dict):
        ''' '''
        self._local_logs_folder = Path(reporter_options.get("logs_folder", "log_records"))
        self._local_errs_folder = Path(reporter_options.get("errs_folder", "log_errors"))
        self._options = reporter_options
        try:
            os.mkdir(self._local_logs_folder)
        except FileExistsError:
            pass
        except Exception as e:
            _top_logger.error(f"Fail to create logs folder with exception {e}")
        try:
            os.mkdir(self._local_errs_folder)
        except FileExistsError:
            pass
        except Exception as e:
            _top_logger.error(f"Fail to create log errors folder with exception {e}")

    def add(self, record:LogRecord):
        ''' '''
        folder = self._local_errs_folder if record.logType == LogRecordType.ERROR else self._local_logs_folder
        with open(folder / f"{str(uuid4())}.json", "w") as f:
            json.dump(record.as_dict(), f)

    def add_bunch(self, records:List[LogRecord]):
        ''' '''

    def _get_source_folder(self, record_type:LogRecordType)->Path:
        source_folder:Union[Path, None] = None
        match record_type:
            case LogRecordType.LATENCY:
                source_folder = self._local_logs_folder
            case LogRecordType.ERROR:
                source_folder = self._local_logs_folder
                        
        if source_folder is None:
            raise ValueError(f"Only LATENCY and ERROR types supported for now")
        return source_folder

    def list_all(self, record_type:LogRecordType)->List[str]:
        ''' '''
        return list(str(self._get_source_folder(record_type).glob("*.json")))

    def get_all(self, record_type:LogRecordType)->List[LogRecord]:
        ''' '''
        result = []
        for one_rec_path in self._get_source_folder(record_type).glob("*.json"):
            with open(one_rec_path, "r") as f:
                try:
                    rec = json.load(f)
                    result.append(LogRecord(**rec))
                except Exception as e:
                    _top_logger.error(f"Fail to read record {one_rec_path} with exception {e}")
        return result

    def get_one(self, record_id)->LogRecord:
        ''' '''
        one_rec_path = record_id if isinstance(record_id, Path) else Path(record_id)
        with open(one_rec_path, "r") as f:
            res_record = LogRecord(**json.load(f))
        return res_record

class ReportAggregator(ABC):
    ''' will use previously created atomic records to create large aggregated report file '''
    @abstractmethod
    def __init__(self, report_aggregator_options:dict={}):
        ''' '''

    @abstractmethod
    def aggregate(self, source:Reporter, record_type:LogRecordType, destination, options:dict):
        ''' '''
    # we're adding some methods to our Abstract class
    # this makes class not 100% pure but more convenient for usage
    @staticmethod
    def _collect_column_names(source:Reporter, record_type:LogRecordType, options:dict={})->Tuple[List[str], List[LogRecord]]:
        ''' 
        supported options keys:

            level_key_separator - default "||=>"

            split_task_value - default '-'
        '''
        _separator = options.get("level_key_separator", "||=>")
        split_task_value = options.get("split_task_value", '-')

        all_records = source.get_all(record_type)
        # we'll collect only fields with basic values and second level field values
        columns = set()
        base_columns:List[str] = ["stage", "job", "task"]
        if isinstance(split_task_value,str):
            base_columns.extend(["task_auth", "task_lang", "task_func", "task_size"])
        for one_rec in all_records:
            if isinstance(one_rec.data, dict) and len(one_rec.data)>0:
                top_keys = set([f"{k}" for k,v in one_rec.data.items() if not isinstance(v,(list,dict,tuple,set))])
                sec_level_keys = set([
                    f"{k}{_separator}{sk}" 
                    for k,v in one_rec.data.items() if isinstance(v,dict)
                        for sk,sv in v.items() if not isinstance(sv,(list,dict,tuple,set))
                ])  
                columns = columns.union(top_keys).union(sec_level_keys)
        columns = columns.difference(base_columns)
        base_columns.extend(list(columns) if len(columns)>0 else [])

        return (base_columns, all_records)
    
class ReportAggregatorCsv(ReportAggregator):
    ''' '''

    def __init__(self, report_aggregator_options:dict={}):
        ''' '''
        self._report_aggregator_options = report_aggregator_options

    
    def aggregate(self, source:Reporter, record_type:LogRecordType, destination:Path, options:dict={}):
        ''' 
        supported options keys:

            level_key_separator - default "||=>"

            comma_replacement - default ';'

            max_value_length - default 40

            split_task_value - default '-'

            timestamp_delta_suffix - default '_delta'
        '''
        self._separator = options.get("level_key_separator", "||=>")
        comma_replacement = options.get("comma_replacement", ";")
        max_value_length = int(options.get("max_value_length", 40))
        split_task_value = options.get("split_task_value", '-')
        timestamp_delta_suffix = options.get("timestamp_delta_suffix", '-')

        # collect column names and log records
        base_columns, all_records = ReportAggregator._collect_column_names(source, record_type, options)
        timestamp_columns = [v for v in base_columns if "timestamp" in v] if isinstance(timestamp_delta_suffix, str) else None
        timestamp_columns = None if timestamp_columns is None or (isinstance(timestamp_columns, list) and len(timestamp_columns)==0) else timestamp_columns
        timestamp_values = { k:[] for k in timestamp_columns} if isinstance(timestamp_columns, list) else None

        # all_records = source.get_all(record_type)
        # # we'll collect only fields with basic values and second level field values
        # columns = set()
        # base_columns:List[str] = ["stage", "job", "task"]
        # if isinstance(split_task_value,str):
        #     base_columns.extend(["task_auth", "task_lang", "task_func", "task_size"])
        # for one_rec in all_records:
        #     if isinstance(one_rec.data, dict) and len(one_rec.data)>0:
        #         top_keys = set([f"{k}" for k,v in one_rec.data.items() if not isinstance(v,(list,dict,tuple,set))])
        #         sec_level_keys = set([
        #             f"{k}{self._separator}{sk}" 
        #             for k,v in one_rec.data.items() if isinstance(v,dict)
        #                 for sk,sv in v.items() if not isinstance(sv,(list,dict,tuple,set))
        #         ])  
        #         columns = columns.union(top_keys).union(sec_level_keys)
        # columns = columns.difference(base_columns)
        # base_columns.extend(list(columns) if len(columns)>0 else [])
        # we have full columns list and can build csv in memory now

        csv_list = [ ','.join(base_columns) ]
        # fill up csv_list using base_columns
        for one_rec in all_records:
            csv_line:List[str] = []
            rec_all_dict = one_rec.as_dict()
            for column_key in base_columns:
                rec_value = rec_all_dict.get(column_key, None)
                if rec_value is None:
                    if column_key.lower().startswith("task_"):
                        # filled already!
                        continue
                    # lookup in data using multi-level key
                    column_key_parts = column_key.split(self._separator)
                    rec_value = one_rec.data if isinstance(one_rec.data, dict) else None
                    for key_part in column_key_parts:
                        rec_value = (rec_value or {}).get(key_part, None)
                    rec_value = None if isinstance(rec_value, (list,dict,tuple,set)) else rec_value
                elif isinstance(split_task_value,str):
                    if column_key.lower()=="task":
                        sub_values = rec_value.split(split_task_value)
                        rec_value = [rec_value]
                        _rec_value = "N/A"
                        for i in range(4):
                            _rec_value = sub_values[i] if i<len(sub_values) else _rec_value
                            rec_value.append(_rec_value)
                if isinstance(rec_value, list):
                    csv_line.extend(rec_value)
                else:
                    if isinstance(timestamp_columns, list) and column_key in timestamp_columns and isinstance(timestamp_values, list):
                        timestamp_values[column_key].append(rec_value)
                    csv_line.append(f"{rec_value or 'N/A'}".replace(",",comma_replacement)[:max_value_length])
            csv_list.append(",".join(csv_line))
        
        
        # save collected report
        with open(destination, "w") as f:
            f.write('\n'.join(csv_list))

class ReportAggregatorXlsx(ReportAggregator):
    ''' '''

    def __init__(self, report_aggregator_options:dict={}):
        ''' '''
        self._report_aggregator_options = report_aggregator_options

    
    def aggregate(self, source:Reporter, record_type:LogRecordType, destination:Path, options:dict={}):
        ''' 
        supported options keys:

            level_key_separator - default "||=>"

            max_value_length - default 40

            split_task_value - default '-'

            timestamp_delta_suffix - default '_delta'
        '''
        self._separator = options.get("level_key_separator", "||=>")
        max_value_length = int(options.get("max_value_length", 40))
        split_task_value = options.get("split_task_value", '-')
        timestamp_delta_suffix = options.get("timestamp_delta_suffix", '_delta')

        # collect column names and log records
        base_columns, all_records = ReportAggregator._collect_column_names(source, record_type, options)
        timestamp_columns = [v for v in base_columns if "timestamp" in v] if isinstance(timestamp_delta_suffix, str) else None
        timestamp_columns = None if timestamp_columns is None or (isinstance(timestamp_columns, list) and len(timestamp_columns)==0) else timestamp_columns
        timestamp_values = { k:{} for k in timestamp_columns} if isinstance(timestamp_columns, list) else None
        # we have full columns list and can build table rows in memory now
        rows_list:Sequence[ Sequence[ Union[str, int, float] ] ] = [ base_columns ]
        # fill up rows_list using base_columns
        for rec_i,one_rec in enumerate(all_records):
            one_row:list = []
            rec_all_dict = one_rec.as_dict()
            for column_key in base_columns:
                rec_value = rec_all_dict.get(column_key, None)
                if rec_value is None:
                    if column_key.lower().startswith("task_"):
                        # filled already!
                        continue
                    # lookup in data using multi-level key
                    column_key_parts = column_key.split(self._separator)
                    rec_value = one_rec.data if isinstance(one_rec.data, dict) else None
                    for key_part in column_key_parts:
                        rec_value = (rec_value or {}).get(key_part, None)
                    rec_value = None if isinstance(rec_value, (list,dict,tuple,set)) else rec_value
                elif isinstance(split_task_value,str):
                    if column_key.lower()=="task":
                        sub_values = rec_value.split(split_task_value)
                        rec_value = [rec_value]
                        _rec_value = "N/A"
                        for i in range(4):
                            _rec_value = sub_values[i] if i<len(sub_values) else _rec_value
                            rec_value.append(_rec_value)
                if isinstance(rec_value, list):
                    one_row.extend(rec_value)
                elif isinstance(rec_value, str) or rec_value is None:
                    one_row.append(f"{rec_value or 'N/A'}"[:max_value_length])
                else:
                    if isinstance(timestamp_columns, list) and column_key in timestamp_columns and isinstance(timestamp_values, dict):
                        timestamp_values[column_key].setdefault(rec_i, rec_value)
                    one_row.append(rec_value)
            rows_list.append(one_row)
        if isinstance(timestamp_columns, list) and isinstance(timestamp_values, dict):
            # we need to add columns for every row
            rows_list[0].extend([f"{v}{timestamp_delta_suffix}" for v in timestamp_columns])
            timestamp_values_min = { k:min(v.values()) for k,v in timestamp_values.items() }
            for i,row in enumerate(rows_list):
                if i==0:
                    continue
                for ts_column in timestamp_columns:
                    # note i-1 for timestamp_values[ts_column] - that's because of title row
                    delta_v = timestamp_values[ts_column][i-1]-timestamp_values_min[ts_column] if i in timestamp_values[ts_column] else ''
                    row.append(delta_v)
        # save collected report as a separate list in the Excel workbook
        # open or create a workbook
        if destination.is_file():
            # load workbook
            wb = openpyxl.load_workbook(
                destination, read_only=False,
                # keep_vba=True, data_only=False,
                # keep_links=True,rich_text=True
            )
        else:
            # create workbook
            wb = openpyxl.Workbook()

        # create a sheet for collected data
        sheet_name = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        wb.create_sheet(title=sheet_name)
        ws = wb.get_sheet_by_name(sheet_name)
        # write down rows
        for row in rows_list:
            ws.append(row)

        # save resulting workbook
        wb.save(destination)